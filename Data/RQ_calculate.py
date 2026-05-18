from openpyxl import load_workbook
from sklearn.metrics import cohen_kappa_score
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
import re
import json
import plotly.graph_objects as go

def RQ2(path):
    all_result = set()
    decrease_result = set()
    increase_result = set()
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["result"]  
    

    for row in ws.iter_rows(min_row=2, values_only=True):
        commit_number = row[2]
        change_direction = row[6]
        UI_change_only = row[11]
        
        if UI_change_only == 0:
            continue
        
        if commit_number is None:
            continue

        if change_direction == "increased":
            increase_result.add(commit_number)
        elif change_direction == "decreased":
            decrease_result.add(commit_number)

        all_result.add(commit_number)

    unique_all_result = list({s.split('-')[1] for s in all_result})
    unique_increase_result = list({s.split('-')[1] for s in increase_result})
    unique_decrease_result = list({s.split('-')[1] for s in decrease_result})

    print("RQ2:")
    print("Decreased:" + str(len(unique_decrease_result)))
    print(unique_decrease_result)
    print("Increased:" + str(len(unique_increase_result)))
    print(unique_increase_result)
    print("Increased or Decreased" + str(len(unique_all_result)))
    print(unique_all_result)
    

def RQ3(path, name):
    categories = ["Text Expansion","Text Contraction",
        "Upward Shift", "Downward Shift", "Leftward Shift", "Rightward Shift", "Height Expansion", "Height Contraction", "Width Expansion", "Width Contraction",
        "Component Removal", "Component Addition", "Other"
    ]

    aggregated_data = {}
    
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["result"]  
    
    for row in ws.iter_rows(values_only=True):
        test_name = row[0]
        effort = row[1]  
        commit_number = row[2]
        change = row[6]
        UI_change_only = row[11]
        
        if UI_change_only == 0:
            continue
        
        if effort != "cursor_travel_distance":
            continue
        
        key = (commit_number, test_name)
        
        if key not in aggregated_data:
            aggregated_data[key] = {
                "direction": change,
                "code_changes": set(),
                "target_changes": set()
            }
        
        if name=="glados":
            for cc in [row[15], row[16], row[18], row[19]]:
                if cc in categories:
                    aggregated_data[key]["code_changes"].add(cc)
            for tc in [row[12], row[13]]:
                if tc in categories:
                    aggregated_data[key]["target_changes"].add(tc)
        elif name=="timeoff":
            for cc in [row[16], row[17], row[19], row[20], row[22], row[23]]:
                if cc in categories:
                    aggregated_data[key]["code_changes"].add(cc)
            for tc in [row[12], row[13], row[14]]:
                if tc in categories:
                    aggregated_data[key]["target_changes"].add(tc)

    flow_code_target = {}
    flow_target_dir = {}

    for key, content in aggregated_data.items():
        if "Other" in content["code_changes"]:
            continue
        
        dir_val = "Increased" if content["direction"] == "increased" else "Decreased"
        
        c_len = len(content["code_changes"])
        t_len = len(content["target_changes"])
        
        if c_len == 0 or t_len == 0:
            continue
            
        weight = 1.0 / (c_len * t_len)
        
        for cc in content["code_changes"]:
            cc_node = f"Code: {cc}"
            for tc in content["target_changes"]:
                tc_node = f"Target: {tc}"
                
                link_ct = (cc_node, tc_node)
                flow_code_target[link_ct] = flow_code_target.get(link_ct, 0.0) + weight
                
                link_td = (tc_node, dir_val)
                flow_target_dir[link_td] = flow_target_dir.get(link_td, 0.0) + weight

    all_nodes_set = set()
    for (src, tgt) in flow_code_target.keys():
        all_nodes_set.add(src)
        all_nodes_set.add(tgt)
    for (src, tgt) in flow_target_dir.keys():
        all_nodes_set.add(src)
        all_nodes_set.add(tgt)

    all_nodes = list(all_nodes_set)
    node_idx = {name: i for i, name in enumerate(all_nodes)}

    sources = []
    targets = []
    values = []

    for (src, tgt), val in flow_code_target.items():
        sources.append(node_idx[src])
        targets.append(node_idx[tgt])
        values.append(val)

    for (src, tgt), val in flow_target_dir.items():
        sources.append(node_idx[src])
        targets.append(node_idx[tgt])
        values.append(val)

    node_totals = [0] * len(all_nodes)
    for t, val in zip(targets, values):
        node_totals[t] += val
    for s, val in zip(sources, values):
        if all_nodes[s].startswith("Code:"):
            node_totals[s] += val

    labeled_nodes = []
    for i, name in enumerate(all_nodes):
        clean_name = name.replace("Code: ", "").replace("Target: ", "")
        if node_totals[i] > 0:
            val = round(node_totals[i], 1)
            if val.is_integer():
                val = int(val)
            labeled_nodes.append(f"{clean_name} ({val})")
        else:
            labeled_nodes.append(clean_name)
            
    node_colors = []
    for name in all_nodes:
        if name.startswith("Code:"):
            node_colors.append("#636EFA")  
        elif name.startswith("Target:"):
            node_colors.append("#EF553B")  
        elif "Increased" in name:
            node_colors.append("#E11D48")  
        elif "Decreased" in name:
            node_colors.append("#11CAA0")  
        else:
            node_colors.append("#AB63FA")  
            
    link_colors = []
    for tgt_idx in targets:
        tgt_name = all_nodes[tgt_idx]
        if "Increased" in tgt_name or "Target: Downward Shift" in tgt_name or "Target: Rightward Shift" in tgt_name:
            link_colors.append("rgba(225, 29, 72, 0.3)") 
        elif "Decreased" in tgt_name or "Target: Upward Shift" in tgt_name or "Target: Leftward Shift" in tgt_name:
            link_colors.append("rgba(17, 202, 160, 0.3)")
        else:
            link_colors.append("rgba(203, 213, 225, 0.4)")
            
    fig = go.Figure(data=[go.Sankey(
        node=dict(
            pad=15,
            thickness=20,
            line=dict(color="black", width=0.5),
            label=labeled_nodes,  
            color=node_colors
        ),
        link=dict(
            source=sources,
            target=targets,
            value=values,
            color=link_colors  
        )
    )])

    fig.update_layout(
        font_size=24,
        width=1300,  
        margin=dict(l=5, r=5, t=50, b=20),
        annotations=[
            dict(
                x=0.0, y=1.09, xref="paper", yref="paper",
                text="<b>Code-changed UI Element</b>", 
                showarrow=False, xanchor="left", font=dict(size=26)
            ),
            dict(
                x=0.5, y=1.09, xref="paper", yref="paper",
                text="<b>Target UI Elements</b>", 
                showarrow=False, xanchor="center", font=dict(size=26)
            ),
            dict(
                x=1.0, y=1.09, xref="paper", yref="paper",
                text="<b>Impact</b>", 
                showarrow=False, xanchor="right", font=dict(size=26)
            )
        ]
    )
    fig.show()
    #fig.write_image("sankey_diagram.pdf")
    


def kappa():
    # 0. N/A
    # 1. Text Expansion Increase in character count
    # 2. Text Contraction Decrease in character count
    # 3. Upward Shift Vertical translation toward the top of the interface
    # 4. Downward Shift Vertical translation toward the bottom of the interface
    # 5. Leftward Shift Horizontal translation toward the left of the interface
    # 6. Rightward Shift Horizontal translation toward the right of the interface
    # 7. Height Expansion Increase in vertical dimension
    # 8. Height Contraction Decrease in vertical dimension
    # 9. Width Expansion Increase in horizontal dimension
    # 10. Width Contraction Decrease in horizontal dimension
    # 11. Component Removal Deletion of an existing element from the UI structure
    # 12. Component Addition Insertion of a new element into the UI structure

    A_code = [1]*14       + [2]*5 + [5]*9 + [7]*11 + [8]*3 + [9]*11       + [11]* 6 + [12]*37
    B_code = [1]*13 + [4] + [2]*5 + [5]*9 + [7]*11 + [8]*3 + [9]*10 + [8] + [11]* 6 + [12]*37

    print("code-changed UI elements:")
    print(cohen_kappa_score(A_code, B_code))

    A_target = [0] + [0]*3 + [3]*21 + [4]*34 + [5]*14 + [6]*18 + [9] + [12]
    B_target = [9] + [4]*3 + [3]*21 + [4]*34 + [5]*14 + [6]*18 + [9] + [12]

    print("target UI elements:")
    print(cohen_kappa_score(A_target, B_target))

    
def parent_child(path, name):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["result"]  
    
    unique_pairs = set()
    unique_indirect_pairs = set()
    
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None and row[1] == "cursor_travel_distance" and row[11]==1:
            test_name = row[0] 
            commit_number = row[2]
            
            unique_pairs.add((test_name, commit_number))
            
            if name == "timeoff":
                if row[24] == 1:
                    unique_indirect_pairs.add((test_name, commit_number))
            elif name == "glados":
                if row[20] == 1:
                    unique_indirect_pairs.add((test_name, commit_number))
    print(name)
    print("total cases:")
    print(len(unique_pairs))
    if name=="timeoff":
        print("Timeoff: 4 'Other' cases identified")
    print("indirect change cases:")
    print(len(unique_indirect_pairs))
    
    

RQ2("./Glados/glados.xlsx")
RQ2("./Timeoff/timeoff.xlsx")

RQ3("./Glados/glados.xlsx", name="glados")
RQ3("./Timeoff/timeoff.xlsx", name="timeoff")

kappa()
    
parent_child("./Glados/glados.xlsx", "glados")
parent_child("./Timeoff/timeoff.xlsx", "timeoff")
# Timeoff: 4 "Other" cases identified